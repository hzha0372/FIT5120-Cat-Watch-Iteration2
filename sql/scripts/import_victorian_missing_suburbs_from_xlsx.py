"""
Import missing Victorian suburbs from an XLSX sheet.

Rules:
- insert rows into suburb_demographics with ON CONFLICT (postcode, suburb_name) DO NOTHING
- skip suspiciously long locality strings (e.g. malformed concatenated content)
- set household_count and population to 0 for inserted rows when sheet values are empty
"""

import os
from pathlib import Path

import psycopg2
from openpyxl import load_workbook


DB_CONFIG = {
    "host": os.getenv("PGHOST", "localhost"),
    "port": int(os.getenv("PGPORT", "5432")),
    "dbname": os.getenv("PGDATABASE", "echoes_of_earth"),
    "user": os.getenv("PGUSER", "postgres"),
    "password": os.getenv("PGPASSWORD", "P@ssw0rd"),
}

XLSX_PATH = Path(
    os.getenv("MISSING_SUBURBS_XLSX", "/tmp/fit5120_epic4_data/victorian_missing_suburbs.xlsx")
)
MAX_LOCALITY_LENGTH = int(os.getenv("MAX_LOCALITY_LENGTH", "180"))


def clean_text(v):
    return str(v or "").strip()


def clean_postcode(v):
    text = clean_text(v)
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 4:
        return digits[:4]
    return ""


def as_float(v):
    if v is None or clean_text(v) == "":
        return None
    try:
        return float(v)
    except Exception:
        return None


def find_col(header_map, *candidates):
    for name in candidates:
        if name in header_map:
            return header_map[name]
    return None


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"XLSX not found: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("XLSX is empty")

    header = [clean_text(x).lower() for x in rows[0]]
    header_map = {name: idx for idx, name in enumerate(header) if name}

    postcode_idx = find_col(header_map, "postcode", "post code", "postal code")
    locality_idx = find_col(header_map, "locality", "suburb", "suburb_name")
    lga_idx = find_col(header_map, "lga", "lga_name", "local government area")
    lat_idx = find_col(header_map, "centroid_lat", "latitude", "lat")
    lng_idx = find_col(header_map, "centroid_lng", "longitude", "lng", "lon")
    state_idx = find_col(header_map, "state")

    if postcode_idx is None or locality_idx is None:
        raise RuntimeError("XLSX must include postcode and locality/suburb columns")

    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    cur = conn.cursor()

    used_rows = 0
    skipped_rows = 0
    inserted_rows = 0
    duplicate_rows = 0

    try:
        for raw in rows[1:]:
            if raw is None:
                continue

            postcode = clean_postcode(raw[postcode_idx] if postcode_idx < len(raw) else "")
            suburb_name = clean_text(raw[locality_idx] if locality_idx < len(raw) else "")
            if not postcode or not suburb_name:
                continue

            if len(suburb_name) > MAX_LOCALITY_LENGTH:
                skipped_rows += 1
                continue

            lga_name = clean_text(raw[lga_idx] if lga_idx is not None and lga_idx < len(raw) else "")
            state = clean_text(raw[state_idx] if state_idx is not None and state_idx < len(raw) else "") or "VIC"
            lat = as_float(raw[lat_idx] if lat_idx is not None and lat_idx < len(raw) else None)
            lng = as_float(raw[lng_idx] if lng_idx is not None and lng_idx < len(raw) else None)

            used_rows += 1
            cur.execute(
                """
                INSERT INTO suburb_demographics
                  (postcode, suburb_name, centroid_lat, centroid_lng,
                   household_count, population, lga_name, state)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (postcode, suburb_name) DO NOTHING
                """,
                (
                    postcode,
                    suburb_name,
                    lat,
                    lng,
                    0,
                    0,
                    lga_name or None,
                    state[:3],
                ),
            )
            if cur.rowcount == 1:
                inserted_rows += 1
            else:
                duplicate_rows += 1

        conn.commit()
        print(f"used_rows: {used_rows}")
        print(f"skipped_rows: {skipped_rows}")
        print(f"inserted_rows: {inserted_rows}")
        print(f"duplicate_rows: {duplicate_rows}")
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()

